from __future__ import annotations

import hashlib
import re
import unicodedata
import zipfile
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from .models import BudgetAnalysis, BudgetItem, Transfer

NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
      "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
      "p": "http://schemas.openxmlformats.org/package/2006/relationships"}
HEADERS = ["Kód", "Název", "Měrná jednotka (individuální)", "Cena jednotky", "Počet jednotek",
           "Částka celkem", "Potomek", "Úroveň", "Procento", "Kombinace veřejné podpory",
           "Měrná jednotka (přednastavena ŘO)", "Měrná jednotka (z číselníku)"]


def decimal(value) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value).replace("\u00a0", "").replace(" ", "").replace(",", "."))
    except InvalidOperation:
        return None


def _column(ref: str) -> int:
    letters = re.match(r"[A-Z]+", ref).group()
    result = 0
    for char in letters:
        result = result * 26 + ord(char) - 64
    return result - 1


def fallback_rows(data: bytes, sheet_name: str = "Export") -> list[list[object]]:
    with zipfile.ZipFile(BytesIO(data)) as z:
        workbook = ET.fromstring(z.read("xl/workbook.xml"))
        sheet = next((s for s in workbook.findall("m:sheets/m:sheet", NS) if s.get("name") == sheet_name), None)
        if sheet is None:
            raise ValueError("V souboru chybí list Export.")
        rid = sheet.get(f"{{{NS['r']}}}id")
        rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        rel = next(x for x in rels if x.get("Id") == rid)
        target = rel.get("Target").lstrip("/")
        path = target if target.startswith("xl/") else "xl/" + target
        root = ET.fromstring(z.read(path))
        shared: list[str] = []
        if "xl/sharedStrings.xml" in z.namelist():
            ss = ET.fromstring(z.read("xl/sharedStrings.xml"))
            shared = ["".join(t.text or "" for t in si.findall(".//m:t", NS)) for si in ss]
        rows: list[list[object]] = []
        for row in root.findall(".//m:sheetData/m:row", NS):
            values: list[object] = []
            for cell in row.findall("m:c", NS):
                index = _column(cell.get("r"))
                while len(values) <= index:
                    values.append(None)
                typ = cell.get("t")
                v = cell.find("m:v", NS)
                inline = cell.find("m:is", NS)
                raw = v.text if v is not None else None
                if inline is not None:
                    value = "".join(t.text or "" for t in inline.findall(".//m:t", NS))
                elif typ == "s":
                    # Vadný export MS2021+ někdy označí inline text jako shared string bez souboru.
                    value = shared[int(raw)] if shared and raw and int(raw) < len(shared) else raw
                elif typ in {"str", "inlineStr"}:
                    value = raw or ""
                elif typ == "b":
                    value = raw == "1"
                elif raw is not None:
                    value = Decimal(raw)
                else:
                    value = None
                values[index] = value
            rows.append(values)
        return rows


def _standard_rows(data: bytes) -> list[list[object]]:
    wb = openpyxl.load_workbook(BytesIO(data), data_only=True, read_only=True)
    if "Export" not in wb.sheetnames:
        raise ValueError("V souboru chybí list Export.")
    return [list(row) for row in wb["Export"].iter_rows(values_only=True)]


def _plain(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    return " ".join("".join(char for char in text if not unicodedata.combining(char)).lower().split())


def parse_financial_plan(source: str | Path | bytes, file_name: str | None = None) -> dict:
    """Parse the IS KP21+ Financial plan export without trusting its calculated totals."""
    data = source if isinstance(source, bytes) else Path(source).read_bytes()
    name = file_name or (Path(source).name if not isinstance(source, bytes) else "financni-plan.xlsx")
    try:
        rows = _standard_rows(data)
    except Exception:
        rows = fallback_rows(data)
    header_index = next((index for index, row in enumerate(rows)
                         if any("poradi financniho planu" in _plain(cell) for cell in row)), None)
    if header_index is None:
        raise ValueError("Soubor není export Finančního plánu z IS KP21+.")
    headers = {_plain(value): index for index, value in enumerate(rows[header_index])}
    required = {
        "order": "poradi financniho planu",
        "coverage": "castka na kryti vydaju - skutecnost",
        "settlement": "vyuctovani - skutecnost",
        "state": "stav zopl",
        "advance": "zalohova platba",
        "final": "zaverecna platba",
    }
    columns: dict[str, int] = {}
    for key, wanted in required.items():
        match = next((index for header, index in headers.items() if wanted in header), None)
        if match is None:
            raise ValueError(f"Ve Finančním plánu chybí sloupec: {wanted}.")
        columns[key] = match
    parsed = []
    for row in rows[header_index + 1:]:
        order = decimal(row[columns["order"]] if len(row) > columns["order"] else None)
        if order is None:
            continue
        parsed.append({
            "sequence_number": int(order),
            "coverage_actual": decimal(row[columns["coverage"]] if len(row) > columns["coverage"] else None) or Decimal("0"),
            "settlement_actual": decimal(row[columns["settlement"]] if len(row) > columns["settlement"] else None) or Decimal("0"),
            "state": str(row[columns["state"]] or "") if len(row) > columns["state"] else "",
            "is_advance_payment": bool(row[columns["advance"]]) if len(row) > columns["advance"] else False,
            "is_final_payment": bool(row[columns["final"]]) if len(row) > columns["final"] else False,
        })
    if not parsed:
        raise ValueError("Finanční plán neobsahuje žádné řádky ŽoP.")
    return {"file_name": name, "sha256": hashlib.sha256(data).hexdigest(), "rows": parsed}


def parse_budget(source: str | Path | bytes, file_name: str | None = None) -> BudgetAnalysis:
    data = source if isinstance(source, bytes) else Path(source).read_bytes()
    name = file_name or (Path(source).name if not isinstance(source, bytes) else "rozpocet.xlsx")
    try:
        rows = _standard_rows(data)
    except Exception:
        rows = fallback_rows(data)
    if not rows:
        raise ValueError("Rozpočet neobsahuje žádné řádky.")
    header_index = next((i for i, r in enumerate(rows) if str(r[0]).strip() == "Kód"), None)
    if header_index is None:
        raise ValueError("V souboru chybí povinné sloupce rozpočtu.")
    items: list[BudgetItem] = []
    warnings: list[str] = []
    seen: set[str] = set()
    for row_no, row in enumerate(rows[header_index + 1:], header_index + 2):
        row = row + [None] * (12-len(row))
        code = str(row[0] or "").strip()
        name_value = str(row[1] or "").strip()
        if not code and not name_value:
            continue
        if not code or not name_value:
            warnings.append(f"Řádek {row_no}: chybí kód nebo název.")
            continue
        if code in seen:
            warnings.append(f"Duplicitní kód {code} na řádku {row_no}.")
        seen.add(code)
        level_value = decimal(row[7])
        level = int(level_value) if level_value is not None else len(code.split("."))
        parent = ".".join(code.split(".")[:-1]) or None
        pct = decimal(row[8])
        normalized_name = name_value.lower()
        ascii_name = "".join(char for char in unicodedata.normalize("NFKD", normalized_name)
                             if not unicodedata.combining(char))
        if code == "2" or code.startswith("2."):
            category = "ineligible"
        elif code in {"3", "4"} or code.startswith(("3.", "4.")):
            category = "informational"
        elif "pauš" in normalized_name or "pausal" in ascii_name or "neprim" in ascii_name:
            category = "lump_sum"
        else:
            category = "direct"
        total_amount = (decimal(row[5]) or Decimal("0")).quantize(Decimal("0.01"))
        items.append(BudgetItem(code=code, name=name_value, parent_code=parent, level=level,
            unit_custom=str(row[2]) if row[2] is not None else None, unit_price=decimal(row[3]),
            unit_count=decimal(row[4]), total_amount=total_amount, percentage=pct,
            support_combination=str(row[9]) if row[9] is not None else None,
            unit_preset=str(row[10]) if row[10] is not None else None,
            unit_catalog=str(row[11]) if row[11] is not None else None, category=category,
            source_row_number=row_no))
    codes = {x.code for x in items}
    for item in items:
        item.is_leaf = not any(x.parent_code == item.code for x in items)
        if item.parent_code and item.parent_code not in codes:
            item.parent_code = next((".".join(item.code.split(".")[:n]) for n in range(len(item.code.split("."))-1, 0, -1)
                                     if ".".join(item.code.split(".")[:n]) in codes), None)
    root = next((x for x in items if x.code == "1"), None)
    lump = next((x for x in items if x.category == "lump_sum"), None)
    rate = (lump.percentage / 100 if lump and lump.percentage and lump.percentage > 1 else lump.percentage) if lump else None
    base_code = None
    if lump and rate:
        base_code = next((x.code for x in items if abs(x.total_amount * rate - lump.total_amount) <= Decimal("0.01")), None)
    return BudgetAnalysis(sha256=hashlib.sha256(data).hexdigest(), file_name=name, items=items,
        total_amount=root.total_amount if root else sum(x.total_amount for x in items if x.parent_code is None),
        lump_sum_rate=rate, lump_sum_base_code=base_code, leaf_count=sum(x.is_leaf for x in items),
        summary_count=sum(not x.is_leaf for x in items), warnings=warnings)


def validate_budget_structure(analysis: BudgetAnalysis) -> list[str]:
    """Check arithmetic that can be broken by manually editing an exported XLSX."""
    errors: list[str] = []
    codes: dict[str, list[BudgetItem]] = {}
    for item in analysis.items:
        codes.setdefault(item.code, []).append(item)
        if item.total_amount < 0:
            errors.append(f"Položka {item.code} nesmí mít zápornou částku.")

    for code, occurrences in codes.items():
        if len(occurrences) > 1:
            rows = ", ".join(str(item.source_row_number) for item in occurrences)
            errors.append(f"Kód {code} je v rozpočtu uveden vícekrát (řádky {rows}).")

    for item in analysis.items:
        children = [child for child in analysis.items if child.parent_code == item.code]
        is_direct_cost_tree = item.code in {"1", "1.1"} or item.code.startswith("1.1.")
        if not children or not is_direct_cost_tree or item.category in {"lump_sum", "informational"}:
            continue
        children_total = sum((child.total_amount for child in children), Decimal("0"))
        if children_total != item.total_amount:
            difference = children_total - item.total_amount
            errors.append(
                f"Součet podřízených položek kódu {item.code} je {children_total:.2f} Kč, "
                f"ale nadřazená položka uvádí {item.total_amount:.2f} Kč "
                f"(rozdíl {difference:+.2f} Kč)."
            )
    return errors


def export_with_formulas(analysis: BudgetAnalysis) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    row_by_code: dict[str, int] = {}
    for item in analysis.items:
        ws.append([item.code, item.name, item.unit_custom, item.unit_price, item.unit_count, item.total_amount,
                   None, item.level, item.percentage, item.support_combination, item.unit_preset, item.unit_catalog])
        row_by_code[item.code] = ws.max_row
    for item in analysis.items:
        row = row_by_code[item.code]
        children = [x for x in analysis.items if x.parent_code == item.code]
        if item.category == "lump_sum" and analysis.lump_sum_base_code:
            ws.cell(row, 6, f"=F{row_by_code[analysis.lump_sum_base_code]}*I{row}/100")
        elif children:
            ws.cell(row, 6, "=" + "+".join(f"F{row_by_code[x.code]}" for x in children))
        elif item.unit_price is not None and item.unit_count is not None:
            ws.cell(row, 6, f"=ROUND(D{row}*E{row},2)")
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 55
    for cell in ws["F"][1:]:
        cell.number_format = '#,##0.00 "Kč"'
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def export_budget_status(rows: list[dict], monthly: dict[str, dict[str, Decimal]]) -> bytes:
    """Export the budget overview and add only months that contain saved financial data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Čerpání rozpočtu"
    months = sorted({month for values in monthly.values() for month, amount in values.items() if amount})
    month_labels = []
    month_names = ("leden", "únor", "březen", "duben", "květen", "červen",
                   "červenec", "srpen", "září", "říjen", "listopad", "prosinec")
    for month in months:
        year, month_number, _ = (int(value) for value in month.split("-"))
        month_labels.append(f"{month_names[month_number - 1]} {year}")
    headers = ["Kód", "Položka", "Rozpočet", *month_labels, "Kumulativně", "Zůstatek", "Čerpání"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E3D")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        values = monthly.get(str(row["code"]), {})
        ws.append([row["code"], row["name"], row["total_amount"],
                   *[values.get(month) for month in months], row["cumulative_spent"],
                   row["remaining"], row["spent_percent"] / Decimal("100")])
        excel_row = ws.max_row
        ws.cell(excel_row, 2).alignment = Alignment(indent=max(0, int(row.get("level", 0)) - 1))
        if not row.get("is_leaf", False):
            for cell in ws[excel_row]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="EAF2ED")
        if Decimal(str(row["remaining"])) < 0:
            ws.cell(excel_row, len(headers) - 1).fill = PatternFill("solid", fgColor="F8D7DA")
    money_columns = list(range(3, 4 + len(months) + 2))
    for column in money_columns:
        for cell in ws.iter_cols(min_col=column, max_col=column, min_row=2):
            cell[0].number_format = '#,##0.00 "Kč"'
    for cell in ws.iter_cols(min_col=len(headers), max_col=len(headers), min_row=2):
        cell[0].number_format = "0.0%"
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 52
    for column in range(3, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(column)].width = 17
    ws.sheet_view.showGridLines = False
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def export_final_settlement(project: dict, breakdown: dict) -> bytes:
    """Create an auditable final-settlement workbook with formulas and source ŽoP rows."""
    wb = openpyxl.Workbook()
    summary = wb.active
    summary.title = "Souhrn"
    detail = wb.create_sheet("Žádosti o platbu")

    summary.append(["KONTROLNÍ VÝPOČET ZÁVĚREČNÉHO VYPOŘÁDÁNÍ", None])
    summary.append(["Projekt", project["project_name"]])
    summary.append(["Registrační číslo", project["project_code"]])
    summary.append(["Celkový rozpočet", project["total_budget"]])
    summary.append(["Podíl prostředků poskytovatele", project["public_funding_rate"]])
    summary.append([])
    summary.append(["Úvodní zálohová platba", "=SUMIF('Žádosti o platbu'!C:C,\"Úvodní záloha\",'Žádosti o platbu'!L:L)"])
    summary.append(["Dosud přijaté platby", "=SUM('Žádosti o platbu'!L:L)"])
    summary.append(["Skutečně schválené způsobilé výdaje", "=SUM('Žádosti o platbu'!M:M)"])
    summary.append(["Předloženo a dosud neschváleno", "=SUM('Žádosti o platbu'!N:N)-B9"])
    summary.append(["Předpokládané způsobilé výdaje po schválení", "=SUM('Žádosti o platbu'!N:N)"])
    summary.append(["Nárok poskytovatele po schválení", "=ROUND((B11-SUM('Žádosti o platbu'!K:K))*B5,2)"])
    summary.append(["Předpokládaný rozdíl (+ doplatek / − vratka)", "=B12-B8"])
    summary.append([])
    summary.append(["Poznámka", "Neschválená závěrečná ŽoP je započtena pouze do orientačního scénáře. Konečnou částku stanoví poskytovatel."])

    for cell in summary[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=13)
        cell.fill = PatternFill("solid", fgColor="1F4E3D")
    summary.merge_cells("A1:B1")
    summary.column_dimensions["A"].width = 48
    summary.column_dimensions["B"].width = 72
    for row in range(4, 14):
        summary.cell(row, 2).number_format = '0.00%' if row == 5 else '#,##0.00 "Kč"'
    summary.cell(13, 2).font = Font(bold=True, color="9C0006")
    summary.cell(13, 2).fill = PatternFill("solid", fgColor="FFC7CE")
    summary.cell(15, 2).alignment = Alignment(wrap_text=True)
    summary.freeze_panes = "A7"
    summary.sheet_view.showGridLines = False

    headers = [
        "ŽoP", "Stav", "Typ", "Soubor", "Prokazované přímé", "Prokazovaný paušál",
        "Prokazováno celkem", "Schválené přímé", "Schválený paušál", "Schváleno celkem",
        "Čisté jiné peněžní příjmy", "Částka na krytí výdajů / zálohová platba", "Započteno jako schválené",
        "Započteno v orientačním scénáři", "Vysvětlení", "Částka uvedená v PDF",
        "Skutečně vyplaceno dle Finančního plánu", "Vyúčtování dle Finančního plánu", "Zdroj Finančního plánu",
    ]
    detail.append(headers)
    for cell in detail[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E3D")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for item in breakdown["rows"]:
        detail.append([
            item["sequence_number"], item["status"], item["type"], item["source_file_name"],
            item["declared_direct"], item["declared_lump_sum"], item["declared_total"],
            item["approved_direct"], item["approved_lump_sum"], item["approved_total"], item["clean_other_income"],
            item["received_payment"], item["approved_for_settlement"],
            item["projected_for_settlement"], item["explanation"], item["pdf_public_payment"],
            item["financial_plan_coverage_actual"], item["financial_plan_settlement_actual"],
            item["financial_plan_source_file_name"],
        ])
        row = detail.max_row
        if not item["is_approved"] and not item["is_advance_payment"]:
            for cell in detail[row]:
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
        detail.cell(row, 15).alignment = Alignment(wrap_text=True, vertical="top")
    for column in list(range(5, 15)) + list(range(16, 19)):
        for cells in detail.iter_cols(min_col=column, max_col=column, min_row=2):
            cells[0].number_format = '#,##0.00 "Kč"'
    widths = {"A": 8, "B": 30, "C": 18, "D": 24, "O": 58}
    for column, width in widths.items():
        detail.column_dimensions[column].width = width
    for column in range(5, 15):
        detail.column_dimensions[openpyxl.utils.get_column_letter(column)].width = 19
    detail.freeze_panes = "E2"
    detail.auto_filter.ref = detail.dimensions
    detail.sheet_view.showGridLines = False

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def export_transfer_proposal(analysis: BudgetAnalysis, transfers: list[Transfer]) -> bytes:
    """Doplní k rozpočtu navržené přesuny a nové kontrolní částky."""
    wb = openpyxl.load_workbook(BytesIO(export_with_formulas(analysis)))
    ws = wb["Export"]
    note_col, count_col, price_col, proposed_col, check_col = 13, 14, 15, 16, 17
    ws.cell(1, note_col, "Navrhovaná změna")
    ws.cell(1, count_col, "Navrhovaný počet jednotek")
    ws.cell(1, price_col, "Navrhovaná cena za jednotku")
    ws.cell(1, proposed_col, "Navrhovaná částka")
    ws.cell(1, check_col, "Kontrola")
    for cell in (ws.cell(1, note_col), ws.cell(1, count_col), ws.cell(1, price_col),
                 ws.cell(1, proposed_col), ws.cell(1, check_col)):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    row_by_code = {str(ws.cell(row, 1).value): row for row in range(2, ws.max_row + 1)}
    incoming: dict[str, list[Transfer]] = {}
    outgoing: dict[str, list[Transfer]] = {}
    for transfer in transfers:
        incoming.setdefault(transfer.target_code, []).append(transfer)
        outgoing.setdefault(transfer.source_code, []).append(transfer)
    for item in analysis.items:
        row = row_by_code[item.code]
        notes = [f"Přijmout {t.amount:.2f} Kč z {t.source_code}" for t in incoming.get(item.code, [])]
        notes += [f"Přesunout {t.amount:.2f} Kč do {t.target_code}" for t in outgoing.get(item.code, [])]
        ws.cell(row, note_col, "; ".join(notes))
        delta = sum((t.amount for t in incoming.get(item.code, [])), Decimal("0")) - sum(
            (t.amount for t in outgoing.get(item.code, [])), Decimal("0"))
        children = [x for x in analysis.items if x.parent_code == item.code]
        ws.cell(row, count_col, f"=E{row}" if item.unit_count is not None else "")
        if item.category == "lump_sum" and analysis.lump_sum_base_code:
            ws.cell(row, proposed_col, f"=P{row_by_code[analysis.lump_sum_base_code]}*I{row}/100")
        elif children:
            ws.cell(row, proposed_col, "=" + "+".join(f"P{row_by_code[x.code]}" for x in children))
        elif item.unit_count and item.unit_count > 0:
            desired = f"F{row}{delta:+.2f}" if delta else f"F{row}"
            ws.cell(row, price_col, f"=ROUND(({desired})/N{row},2)")
            ws.cell(row, proposed_col, f"=ROUND(N{row}*O{row},2)")
            ws.cell(row, check_col, f'=IF(P{row}=ROUND({desired},2),"OK","NELZE: počet × cena nedá požadovanou částku")')
        elif delta:
            ws.cell(row, proposed_col, f"=F{row}{delta:+.2f}")
        else:
            ws.cell(row, proposed_col, f"=F{row}")
        if not ws.cell(row, check_col).value:
            ws.cell(row, check_col, "OK")
        ws.cell(row, count_col).number_format = '0.00'
        ws.cell(row, price_col).number_format = '#,##0.00 "Kč"'
        ws.cell(row, proposed_col).number_format = '#,##0.00 "Kč"'
    ws.column_dimensions["M"].width = 55
    ws.column_dimensions["N"].width = 24
    ws.column_dimensions["O"].width = 25
    ws.column_dimensions["P"].width = 22
    ws.column_dimensions["Q"].width = 42
    output = BytesIO()
    wb.save(output)
    return output.getvalue()
