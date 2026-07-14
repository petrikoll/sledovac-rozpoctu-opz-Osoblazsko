from decimal import Decimal
from io import BytesIO
from pathlib import Path
import zipfile

import openpyxl

from app.pdf_parser import extract_budget_code, money, parse_payment_request
from app.models import Transfer
from app.xlsx_parser import export_budget_status, export_transfer_proposal, export_with_formulas, fallback_rows, parse_budget, parse_financial_plan, validate_budget_structure

SAMPLES = Path(__file__).parents[1] / "samples"


def test_parse_iskp_financial_plan():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Export"
    sheet.append(["Součtový řádek", "Pořadí finančního plánu", "Datum předložení",
                  "Částka na krytí výdajů - plán", "Vyúčtování - plán",
                  "Částka na krytí výdajů - skutečnost", "Vyúčtování - skutečnost",
                  "Stav ŽoPl", "Zálohová platba", "Závěrečná platba"])
    sheet.append([None, 1, None, 1000, 0, 980, 0, "Proplacená", True, False])
    sheet.append([None, 2, None, 0, 900, 0, 875, "Zaregistrovaná", False, True])
    output = BytesIO()
    workbook.save(output)

    result = parse_financial_plan(output.getvalue(), "financni-plan.xlsx")

    assert result["file_name"] == "financni-plan.xlsx"
    assert result["rows"][0]["coverage_actual"] == Decimal("980")
    assert result["rows"][1]["settlement_actual"] == Decimal("875")
    assert result["rows"][1]["is_final_payment"] is True


def test_budget_code_with_internal_pdf_spaces_is_normalized():
    assert extract_budget_code("1.1.4.1 .1 Nájem prostor") == "1.1.4.1.1"


def test_same_item_name_is_distinguished_by_full_budget_code():
    assert extract_budget_code("1.1.4.2 .1 Právník") == "1.1.4.2.1"
    assert extract_budget_code("1.1.1.3 .3 Právník") == "1.1.1.3.3"


def test_money_with_spaces_around_decimal_digits():
    assert money("6 110, 0 0") == Decimal("6110.00")


def test_real_budget():
    result = parse_budget(SAMPLES / "Export_2026-07-11_084920.xlsx")
    assert result.total_amount == Decimal("4415040")
    assert result.lump_sum_rate == Decimal("0.4")
    assert result.lump_sum_base_code == "1.1"
    values = {x.code: x for x in result.items}
    assert values["1.1"].total_amount == Decimal("3153600")
    assert values["1.2"].total_amount == Decimal("1261440")
    assert values["1.1.1.1"].total_amount == Decimal("2232000")
    assert values["1.1.1.2"].total_amount == Decimal("921600")
    assert values["2"].category == "ineligible"
    assert values["3"].category == "informational"
    assert values["4"].category == "informational"
    assert validate_budget_structure(result) == []


def test_budget_structure_rejects_inconsistent_parent_totals():
    result = parse_budget(SAMPLES / "Export_2026-07-11_084920.xlsx")
    item = next(value for value in result.items if value.code == "1.1.1")
    item.total_amount += Decimal("1")

    errors = validate_budget_structure(result)

    assert any("1.1.1" in error and "rozdíl" in error for error in errors)


def test_fallback_and_formula_export():
    data = (SAMPLES / "Export_2026-07-11_084920.xlsx").read_bytes()
    assert fallback_rows(data)
    result = parse_budget(data)
    wb = openpyxl.load_workbook(BytesIO(export_with_formulas(result)), data_only=False)
    rows = {wb["Export"].cell(r, 1).value: r for r in range(2, wb["Export"].max_row + 1)}
    assert wb["Export"].cell(rows["1.1.1"], 6).value.startswith("=")
    assert wb["Export"].cell(rows["1.2"], 6).value == f"=F{rows['1.1']}*I{rows['1.2']}/100"


def test_budget_status_export_adds_only_filled_months():
    rows = [{
        "code": "1.1.1.1", "name": "Psycholog", "level": 4, "is_leaf": True,
        "total_amount": Decimal("100000"), "cumulative_spent": Decimal("30000"),
        "remaining": Decimal("70000"), "spent_percent": Decimal("30"),
    }]
    data = export_budget_status(rows, {"1.1.1.1": {
        "2026-05-01": Decimal("12000"), "2026-06-01": Decimal("18000"), "2026-07-01": Decimal("0"),
    }})

    wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
    ws = wb["Čerpání rozpočtu"]
    assert [cell.value for cell in ws[1]] == ["Kód", "Položka", "Rozpočet", "květen 2026", "červen 2026", "Kumulativně", "Zůstatek", "Čerpání"]
    assert ws.cell(2, 4).value == 12000
    assert ws.cell(2, 5).value == 18000
    assert ws.cell(2, 6).value == 30000


def test_transfer_proposal_export_contains_audit_columns():
    result = parse_budget(SAMPLES / "Export_2026-07-11_084920.xlsx")
    data = export_transfer_proposal(result, [Transfer(
        source_code="1.1.1.1", target_code="1.1.1.2", amount=Decimal("43477.20"))])
    wb = openpyxl.load_workbook(BytesIO(data), data_only=False)
    ws = wb["Export"]
    rows = {str(ws.cell(r, 1).value): r for r in range(2, ws.max_row + 1)}
    assert ws.cell(1, 13).value == "Navrhovaná změna"
    assert "do 1.1.1.2" in ws.cell(rows["1.1.1.1"], 13).value
    assert "z 1.1.1.1" in ws.cell(rows["1.1.1.2"], 13).value
    assert ws.cell(1, 14).value == "Navrhovaný počet jednotek"
    assert ws.cell(1, 15).value == "Navrhovaná cena za jednotku"
    assert ws.cell(1, 16).value == "Navrhovaná částka"
    assert ws.cell(rows["1.1.1.1"], 16).value == f"=ROUND(N{rows['1.1.1.1']}*O{rows['1.1.1.1']},2)"
    assert "ROUND" in ws.cell(rows["1.1.1.1"], 15).value
    assert "-43477.20" in ws.cell(rows["1.1.1.1"], 15).value
    assert "+43477.20" in ws.cell(rows["1.1.1.2"], 15).value


def test_budget_money_is_normalized_to_cents():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export"
    ws.append(["Kód", "Název", "Měrná jednotka (individuální)", "Cena jednotky", "Počet jednotek",
               "Částka celkem", "Potomek", "Úroveň", "Procento", "Kombinace veřejné podpory",
               "Měrná jednotka (přednastavena ŘO)", "Měrná jednotka (z číselníku)"])
    ws.append(["1", "Celkem", None, 0, 0, Decimal("34221.00"), False, 1])
    ws.append(["1.1", "Položka", None, Decimal("287.33"), Decimal("119.10"), Decimal("34221.003"), False, 2])
    output = BytesIO(); wb.save(output)

    result = parse_budget(output.getvalue())

    assert {item.code: item.total_amount for item in result.items}["1.1"] == Decimal("34221.00")
    assert validate_budget_structure(result) == []


def test_payment_samples():
    p0 = parse_payment_request(SAMPLES / "ZOP_PRJ0.pdf")
    p1 = parse_payment_request(SAMPLES / "ZOP_PRJ1.pdf")
    p2 = parse_payment_request(SAMPLES / "ZOP_PRJ2.pdf")
    assert p0.sequence_number == 1 and p0.is_advance_payment and p0.approved_total == 0
    assert p0.public_payment == Decimal("1258286.40")
    assert (p1.approved_direct_costs, p1.approved_lump_sum) == (Decimal("435105.00"), Decimal("174042.00"))
    assert (p2.approved_direct_costs, p2.approved_lump_sum) == (Decimal("514508.00"), Decimal("205803.20"))
    assert p1.approved_direct_costs + p2.approved_direct_costs == Decimal("949613")
    assert p1.approved_lump_sum + p2.approved_lump_sum == Decimal("379845.20")
    assert p1.approved_total + p2.approved_total == Decimal("1329458.20")
